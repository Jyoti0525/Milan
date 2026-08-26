"""Reconciling a merchant's own files, and proving it changed nothing.

The generated path has an answer key. The imported path cannot: a merchant's
files are the input, and nobody hands over the truth alongside them. So the
question "did the import read this correctly" has no ground truth to appeal
to - which is exactly the gap this test closes.

The trick is to make one. A generated dataset is exported to CSV in the shape
a real export arrives in - grouped rupees, a rupee symbol, `Cr` markers, a
bank's banner above the header, a two-digit year - and then imported back
through the full pipeline. The engine has already reconciled the same month
from its own records, so the two reports can be compared line for line.

If the import drops a paisa, reads a date the other way round, or maps the
debit column onto the credit, the reports diverge and this fails. Every
proof, every strategy, every exception and every amount has to match. That is
a stronger statement than "the import ran": it is that reading a merchant's
files produces the same answer as having the data natively, which is the only
claim that makes the feature worth having.
"""

from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import Workbook

from milan.chaos.config import Difficulty, GenerationConfig
from milan.chaos.generator import ChaosEngine
from milan.domain.dataset import Dataset
from milan.domain.money import Paise
from milan.domain.rates import RateCard
from milan.domain.results import ReconReport
from milan.evaluation.harness import to_recon_input
from milan.ingest import build
from milan.ingest.plan import to_saved
from milan.ingest.resolver import Importer, decisions_from
from milan.ingest.schema import RecordKind
from milan.recon.pipeline import ReconciliationPipeline, RunMetadata

SEED = 9
ORDERS = 120


@pytest.fixture(scope="module")
def dataset() -> Dataset:
    return ChaosEngine(
        GenerationConfig(
            seed=SEED,
            difficulty=Difficulty.REALISTIC,
            order_count=ORDERS,
            span_days=21,
            rates=RateCard(),
        )
    ).generate()


def grouped(paise: Paise) -> str:
    """Rupees in the Indian grouping, the way an export writes them."""
    sign = "-" if paise < 0 else ""
    whole, fraction = divmod(abs(paise), 100)
    digits = str(whole)
    if len(digits) > 3:
        head, tail = digits[:-3], digits[-3:]
        groups: list[str] = []
        while len(head) > 2:
            head, group = head[:-2], head[-2:]
            groups.insert(0, group)
        if head:
            groups.insert(0, head)
        digits = ",".join([*groups, tail])
    return f"{sign}{digits}.{fraction:02d}"


def plain(paise: Paise) -> str:
    sign = "-" if paise < 0 else ""
    whole, fraction = divmod(abs(paise), 100)
    return f"{sign}{whole}.{fraction:02d}"


def export(dataset: Dataset, root: Path) -> Path:
    """Write the dataset out as three files a merchant would actually hold.

    Not the engine's own JSON renamed. The whole point is to go out through
    the formats other people's software produces and come back through the
    reader, so every difference between the two is a defect this test can
    see - the rupee symbol, the thousands separators, the `Cr` marker, the
    banner lines, the two-digit year.
    """
    root.mkdir(parents=True, exist_ok=True)

    with (root / "settlement_recon.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "entity_id",
                "type",
                "amount",
                "credit",
                "debit",
                "fee",
                "tax",
                "created_at",
                "settlement_id",
                "settled_at",
                "settlement_utr",
                "payment_id",
                "order_id",
                "method",
                "card_type",
                "on_hold",
                "settled",
            ]
        )
        for row in dataset.settlement_rows:
            writer.writerow(
                [
                    row.entity_id,
                    row.type.value,
                    plain(row.amount),
                    plain(row.credit),
                    plain(row.debit),
                    plain(row.fee),
                    plain(row.tax),
                    row.created_at.isoformat(sep=" "),
                    row.settlement_id or "",
                    row.settled_at.isoformat(sep=" ") if row.settled_at else "",
                    row.settlement_utr or "",
                    row.payment_id or "",
                    row.order_id or "",
                    row.method.value if row.method else "",
                    row.card_type.value if row.card_type else "",
                    "Y" if row.on_hold else "N",
                    "Y" if row.settled else "N",
                ]
            )

    with (root / "OpTxnHistory.csv").open("w", newline="", encoding="utf-8") as handle:
        handle.write("HDFC BANK LIMITED\n")
        handle.write("Statement of account\n")
        handle.write("Account No :50100XXXXXX1234\n")
        handle.write("\n")
        writer = csv.writer(handle)
        writer.writerow(["Value Date", "Narration", "Ref No", "Withdrawal Amt.", "Credit"])
        for credit in sorted(dataset.bank_credits, key=lambda c: (c.value_date, c.credit_id)):
            writer.writerow(
                [
                    credit.value_date.strftime("%d-%b-%Y"),
                    credit.narration,
                    credit.utr or "",
                    "",
                    f"₹ {grouped(credit.amount)} Cr",
                ]
            )

    with (root / "captures.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["payment_id", "order_id", "amount", "method", "card_type", "captured_at"])
        for payment in dataset.payments:
            writer.writerow(
                [
                    payment.payment_id,
                    payment.order_id,
                    grouped(payment.amount),
                    payment.method.value,
                    payment.card_type.value if payment.card_type else "",
                    payment.captured_at.isoformat(sep=" "),
                ]
            )
    return root


def native(dataset: Dataset) -> ReconReport:
    return ReconciliationPipeline().run(
        to_recon_input(dataset),
        RunMetadata(seed=dataset.seed, difficulty=dataset.difficulty),
    )


@pytest.fixture(scope="module")
def imported(dataset: Dataset, tmp_path_factory: pytest.TempPathFactory) -> ReconReport:
    root = export(dataset, tmp_path_factory.mktemp("merchant"))
    plan = Importer(None).plan(root)
    assert plan.ready, plan.blockers()
    return ReconciliationPipeline().run(
        build.build(plan).data,
        RunMetadata(seed=dataset.seed, difficulty=dataset.difficulty),
    )


class TestAMerchantsOwnFilesReconcileToTheSameAnswer:
    def test_no_model_is_needed_to_read_a_normally_named_export(
        self, dataset: Dataset, tmp_path: Path
    ) -> None:
        """The claim that has to survive the demo going offline. A folder whose
        column names are ones somebody has met before imports with no provider
        configured and no question asked."""
        plan = Importer(None).plan(export(dataset, tmp_path / "merchant"))
        assert plan.consulted == "none"
        assert plan.questions == ()
        assert plan.ready

    def test_every_record_survives_the_trip_through_csv(
        self, dataset: Dataset, tmp_path: Path
    ) -> None:
        plan = Importer(None).plan(export(dataset, tmp_path / "merchant"))
        result = build.build(plan)
        assert result.dropped == ()
        assert result.counts["settlement_rows"] == len(dataset.settlement_rows)
        assert result.counts["bank_credits"] == len(dataset.bank_credits)
        assert result.counts["payments"] == len(dataset.payments)

    def test_the_same_credits_are_proved_by_the_same_rungs(
        self, dataset: Dataset, imported: ReconReport
    ) -> None:
        """A date read the other way round, or a debit read as a credit, moves
        a credit from one rung of the cascade to another long before it changes
        a total. This is the sensitive assertion."""
        expected = native(dataset)
        assert Counter(p.strategy.value for p in imported.proofs) == Counter(
            p.strategy.value for p in expected.proofs
        )

    def test_every_proved_amount_matches_to_the_paisa(
        self, dataset: Dataset, imported: ReconReport
    ) -> None:
        expected = native(dataset)
        assert sorted(p.credit_amount for p in imported.proofs) == sorted(
            p.credit_amount for p in expected.proofs
        )

    def test_the_exception_list_is_the_same_list(
        self, dataset: Dataset, imported: ReconReport
    ) -> None:
        """The honest half of the output, and the half a merchant acts on."""
        expected = native(dataset)
        assert sorted((e.code.value, e.amount) for e in imported.exceptions) == sorted(
            (e.code.value, e.amount) for e in expected.exceptions
        )

    def test_the_shortfalls_are_the_same_shortfalls(
        self, dataset: Dataset, imported: ReconReport
    ) -> None:
        expected = native(dataset)
        assert sorted(s.residual for s in imported.shortfalls) == sorted(
            s.residual for s in expected.shortfalls
        )

    def test_the_leak_findings_survive_too(self, dataset: Dataset, imported: ReconReport) -> None:
        """Leak detection reads the card type off every settlement row, which
        is an optional column. If the import dropped it, the books would still
        balance and the overcharge would vanish silently - which is the exact
        failure mode leak detection exists to catch."""
        expected = native(dataset)
        assert sorted(leak.overcharge for leak in imported.leaks) == sorted(
            leak.overcharge for leak in expected.leaks
        )


class TestTheImportedRunIsReproducible:
    def test_importing_the_same_folder_twice_reads_it_the_same_way(
        self, dataset: Dataset, tmp_path: Path
    ) -> None:
        root = export(dataset, tmp_path / "merchant")
        first = Importer(None).plan(root)
        second = Importer(None).plan(root, decisions_from(to_saved(first)))
        assert second.ready
        assert build.build(first).data == build.build(second).data

    def test_a_saved_mapping_pins_the_date_format_it_settled_on(
        self, dataset: Dataset, tmp_path: Path
    ) -> None:
        """The one decision that would otherwise be free to drift. Two readings
        of the same column produce two different months, and a mapping that
        recorded the column but not the format would reproduce neither."""
        root = export(dataset, tmp_path / "merchant")
        saved = to_saved(Importer(None).plan(root))
        patterns = {
            entry.file: {column.field: column.pattern for column in entry.columns if column.pattern}
            for entry in saved.files
        }
        assert patterns["OpTxnHistory.csv"]["value_date"]


class TestWhatTheFolderDidNotContainIsSaidOutLoud:
    def test_a_folder_with_no_orders_file_reports_the_gap(
        self, dataset: Dataset, tmp_path: Path
    ) -> None:
        """The export writes no orders file, and the run is still complete -
        but a merchant reading a clean exception list has a right to know
        which checks were switched off before they trust it."""
        plan = Importer(None).plan(export(dataset, tmp_path / "merchant"))
        assert plan.all_of(RecordKind.ORDERS) == ()
        assert any("no orders file" in line for line in plan.limitations())

    def test_a_statement_with_an_empty_withdrawal_column_says_what_it_costs(
        self, dataset: Dataset, tmp_path: Path
    ) -> None:
        plan = Importer(None).plan(export(dataset, tmp_path / "merchant"))
        assert any("no debit column" in line for line in plan.limitations())


class TestTheReaderIsNotQuietlyRepairingAnything:
    def test_a_row_that_will_not_read_is_dropped_and_reported_with_its_line(
        self, dataset: Dataset, tmp_path: Path
    ) -> None:
        """Silently repairing a bad row is how a merchant ends up four rupees
        short with nothing to look at."""
        root = export(dataset, tmp_path / "merchant")
        path = root / "captures.csv"
        lines = path.read_text(encoding="utf-8").splitlines()
        lines.insert(3, "pay_broken,order_broken,not-an-amount,upi,,2026-07-06 10:00:00")
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")

        plan = Importer(None).plan(root)
        result = build.build(plan)
        assert len(result.dropped) == 1
        assert result.dropped[0].file == "captures.csv"
        assert result.dropped[0].line == 4
        assert result.counts["payments"] == len(dataset.payments)

    def test_a_statement_line_that_is_a_withdrawal_is_counted_not_dropped(
        self, dataset: Dataset, tmp_path: Path
    ) -> None:
        """Money leaving the account is not an error and not a credit. It has
        to be neither, or the row totals stop adding up."""
        root = export(dataset, tmp_path / "merchant")
        path = root / "OpTxnHistory.csv"
        text = path.read_text(encoding="utf-8")
        text += "12-Jul-2026,RENT PAYMENT,,₹ 45,000.00 Dr,\n"
        path.write_text(text, encoding="utf-8")

        plan = Importer(None).plan(root)
        assert plan.ready, plan.blockers()
        result = build.build(plan)
        assert result.withdrawals == 1
        assert result.dropped == ()
        assert result.counts["bank_credits"] == len(dataset.bank_credits)


def export_workbook(dataset: Dataset, root: Path) -> Path:
    """The same month, as one Excel workbook with three sheets.

    This is the shape the files most often actually arrive in. A gateway's
    dashboard exports a workbook; a bank's net banking offers `.xls` above
    `.csv`; and a finance team that has opened either has saved it back as
    `.xlsx`. A merchant handing over "my settlement report and my statement"
    is frequently handing over one file with several sheets in it.

    Deliberately typed rather than stringified. Dates go in as `datetime` and
    amounts as `float`, because that is what a spreadsheet holds and the
    interesting question is whether the reader gets the merchant's own numbers
    back out of Excel's binary doubles - not whether it can read a CSV that
    has been renamed.
    """
    root.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    # A fresh workbook comes with one empty sheet; the three below are
    # created by name, so the default has to go or the import sees a fourth
    # table with nothing in it.
    del book[str(book.sheetnames[0])]

    settlements = book.create_sheet("Settlements")
    settlements.append(
        [
            "entity_id",
            "type",
            "amount",
            "credit",
            "debit",
            "fee",
            "tax",
            "created_at",
            "settlement_id",
            "settled_at",
            "settlement_utr",
            "payment_id",
            "order_id",
            "method",
            "card_type",
            "on_hold",
            "settled",
        ]
    )
    for row in dataset.settlement_rows:
        settlements.append(
            [
                row.entity_id,
                row.type.value,
                row.amount / 100,
                row.credit / 100,
                row.debit / 100,
                row.fee / 100,
                row.tax / 100,
                row.created_at,
                row.settlement_id or "",
                row.settled_at or "",
                row.settlement_utr or "",
                row.payment_id or "",
                row.order_id or "",
                row.method.value if row.method else "",
                row.card_type.value if row.card_type else "",
                row.on_hold,
                row.settled,
            ]
        )

    statement = book.create_sheet("Bank Statement")
    # The banner an actual downloaded statement carries, in the cells it
    # carries it in. The header finder has to walk past it here exactly as it
    # does in a CSV.
    statement.append(["HDFC BANK LIMITED"])
    statement.append(["Account No :50100XXXXXX1234"])
    statement.append([])
    statement.append(["Value Date", "Narration", "Ref No", "Withdrawal Amt.", "Deposit Amt."])
    for credit in sorted(dataset.bank_credits, key=lambda c: (c.value_date, c.credit_id)):
        statement.append(
            [credit.value_date, credit.narration, credit.utr or "", None, credit.amount / 100]
        )

    captures = book.create_sheet("Payments")
    captures.append(["payment_id", "order_id", "amount", "method", "card_type", "captured_at"])
    for payment in dataset.payments:
        captures.append(
            [
                payment.payment_id,
                payment.order_id,
                payment.amount / 100,
                payment.method.value,
                payment.card_type.value if payment.card_type else "",
                payment.captured_at,
            ]
        )

    book.save(root / "razorpay_july_2026.xlsx")
    return root


@pytest.fixture(scope="module")
def from_workbook(dataset: Dataset, tmp_path_factory: pytest.TempPathFactory) -> ReconReport:
    root = export_workbook(dataset, tmp_path_factory.mktemp("workbook"))
    plan = Importer(None).plan(root)
    assert plan.ready, plan.blockers()
    return ReconciliationPipeline().run(
        build.build(plan).data,
        RunMetadata(seed=dataset.seed, difficulty=dataset.difficulty),
    )


class TestOneWorkbookIsAsGoodAsThreeCsvs:
    """The format most merchants actually have, held to the same standard.

    Every assertion below has a twin in the CSV class above. That is the
    point: supporting spreadsheets is only worth anything if it produces the
    same answer, and "we can open .xlsx files" is a claim about a library
    rather than about this engine.
    """

    def test_three_sheets_become_three_tables(self, dataset: Dataset, tmp_path: Path) -> None:
        """Reading only the active sheet would import a third of the month and
        balance perfectly over it - the one failure here that produces no
        error anywhere."""
        plan = Importer(None).plan(export_workbook(dataset, tmp_path / "wb"))
        assert len(plan.placed) == 3
        assert {mapping.kind for mapping in plan.placed} == {
            RecordKind.SETTLEMENT_ROWS,
            RecordKind.BANK_CREDITS,
            RecordKind.PAYMENTS,
        }

    def test_no_model_is_needed_and_nothing_is_asked(
        self, dataset: Dataset, tmp_path: Path
    ) -> None:
        plan = Importer(None).plan(export_workbook(dataset, tmp_path / "wb"))
        assert plan.consulted == "none"
        assert plan.questions == ()

    def test_every_record_survives_the_trip_through_excel(
        self, dataset: Dataset, tmp_path: Path
    ) -> None:
        """Including the ones whose amounts went in as binary doubles."""
        result = build.build(Importer(None).plan(export_workbook(dataset, tmp_path / "wb")))
        assert result.dropped == ()
        assert result.counts["settlement_rows"] == len(dataset.settlement_rows)
        assert result.counts["bank_credits"] == len(dataset.bank_credits)
        assert result.counts["payments"] == len(dataset.payments)

    def test_not_one_paisa_is_lost_to_a_float(self, dataset: Dataset, tmp_path: Path) -> None:
        """The assertion this whole format hangs on. Excel had every amount as
        a double for the length of the round trip, and the sum has to come back
        exactly."""
        result = build.build(Importer(None).plan(export_workbook(dataset, tmp_path / "wb")))
        assert sum(credit.amount for credit in result.data.bank_credits) == sum(
            credit.amount for credit in dataset.bank_credits
        )
        assert sum(row.credit for row in result.data.settlement_rows) == sum(
            row.credit for row in dataset.settlement_rows
        )

    def test_the_same_credits_are_proved_by_the_same_rungs(
        self, dataset: Dataset, from_workbook: ReconReport
    ) -> None:
        expected = native(dataset)
        assert Counter(p.strategy.value for p in from_workbook.proofs) == Counter(
            p.strategy.value for p in expected.proofs
        )

    def test_the_exception_list_is_the_same_list(
        self, dataset: Dataset, from_workbook: ReconReport
    ) -> None:
        expected = native(dataset)
        assert sorted((e.code.value, e.amount) for e in from_workbook.exceptions) == sorted(
            (e.code.value, e.amount) for e in expected.exceptions
        )

    def test_the_leak_findings_survive_too(
        self, dataset: Dataset, from_workbook: ReconReport
    ) -> None:
        """Card type is an optional column and a boolean-ish enum, which is
        exactly the sort of value a spreadsheet mangles."""
        expected = native(dataset)
        assert sorted(leak.overcharge for leak in from_workbook.leaks) == sorted(
            leak.overcharge for leak in expected.leaks
        )
