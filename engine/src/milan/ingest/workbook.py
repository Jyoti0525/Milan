"""Files that are tables but are not CSV, and files that only look like they are.

CSV was never the whole answer. A merchant asked for "your settlement report
and your bank statement" reaches for whatever the download button gave them,
and in India that button gives an Excel workbook more often than anything
else. Razorpay's dashboard exports `.xlsx`. HDFC and ICICI net banking both
offer `.xls` above `.csv` in the list. A finance team that has touched the
file at all has touched it in Excel, and Excel saves what it opens.

So this module does two jobs, and the second one matters as much as the first.

**It reads workbooks.** An `.xlsx` becomes one table per sheet, because a real
export is not one table - Razorpay's is a `Settlements` sheet and a
`Transactions` sheet in one file, and reading the first and ignoring the rest
would silently drop half of what the merchant handed over.

**It names what it cannot read.** A legacy `.xls`, a PDF statement, an HTML
table that a bank saved under an `.xls` extension - each is refused with the
sentence that gets the person unstuck, rather than with "unsupported format".
Those three cover essentially every file that arrives here and is not a table
we can parse, and the difference between "we cannot read this" and "open it in
Excel and Save As CSV" is the difference between a dead end and thirty
seconds of work.

Detection is by content, not by extension. Banks lie about extensions
constantly - ICICI's "Excel" download is an HTML table called `.xls`, and a
statement saved from a browser is often a `.csv` that is really a web page.
The first few bytes of a file do not lie.
"""

from __future__ import annotations

import datetime as dt
import zipfile
from pathlib import Path
from typing import Any

WORKBOOK_SUFFIXES = frozenset({".xlsx", ".xlsm"})
"""What `openpyxl` reads. `.xls` is a different format entirely - see below."""

TEXT_SUFFIXES = frozenset({".csv", ".tsv", ".txt"})

READABLE = WORKBOOK_SUFFIXES | TEXT_SUFFIXES

MAX_FRACTION_DIGITS = 6
"""How many decimal places survive the trip out of a spreadsheet.

Excel holds every number as a binary double, so a column of rupee amounts
that has been through a `SUM` can arrive as `1234.5600000000001`. That is
Excel's arithmetic, not ours, and it has to become a decimal string exactly
once - here, at the boundary - so that everything downstream is `Decimal` on
a string that says what the merchant sees in their own spreadsheet.

Six rather than two, because this converts rates as well as amounts and a
contracted rate of `0.0215` truncated to two places is a fee card that says
two percent when the merchant is paying 2.15.
"""


class FormatError(RuntimeError):
    """This file is not a table this reader can open.

    Deliberately not `UnreadableFileError`: that lives in `reading`, which
    imports this module, and an exception travelling the other way would make
    the two mutually dependent. `reading` translates.
    """


# --------------------------------------------------------------- diagnosis

_SIGNATURES: tuple[tuple[bytes, str], ...] = (
    (
        b"%PDF",
        "this is a PDF, not a table. A PDF statement has no columns to read - "
        "it has ink in the shape of columns. Download the same statement as "
        "CSV or Excel from your bank instead; every major Indian bank offers "
        "both next to the PDF.",
    ),
    (
        b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1",
        "this is a legacy Excel file (.xls, the format Excel used before "
        "2007). Open it in Excel or LibreOffice and use Save As to write it "
        "as .xlsx or .csv, and it will read.",
    ),
    (
        b"PK\x03\x04",
        "",  # A zip. Could be a real .xlsx - resolved by suffix below.
    ),
)

_HTML_MARKERS: tuple[bytes, ...] = (b"<html", b"<!doctype html", b"<table", b"<?xml version")


def diagnose(path: Path) -> str:
    """Why this file cannot be read as a table, or an empty string if it can.

    Checked before anything tries to decode the file as text, because every
    format named here decodes as text perfectly well and then produces a
    header row of gibberish rather than an error. A PDF read as CSV does not
    fail; it succeeds, and yields one column called `%PDF-1.4` - which is a
    far worse outcome than a refusal.
    """
    try:
        with path.open("rb") as handle:
            head = handle.read(1024)
    except OSError as failure:  # pragma: no cover - unreadable on disk
        return f"could not be opened: {failure}"

    if not head.strip():
        return "is empty"

    suffix = path.suffix.lower()

    for signature, reason in _SIGNATURES:
        if not head.startswith(signature):
            continue
        if signature == b"PK\x03\x04":
            # Every `.xlsx` is a zip, but so is every `.docx`, `.odt` and
            # `.zip`. The suffix is the only thing that separates them and
            # here it is being trusted on purpose - a zip whose name says
            # spreadsheet is opened as one, and openpyxl says so if it is not.
            if suffix in WORKBOOK_SUFFIXES:
                return ""
            return (
                "this is a zipped document rather than a table. If it is a "
                "spreadsheet, save it as .xlsx or .csv; if it is an archive, "
                "unzip it first and hand over the files inside."
            )
        return reason

    lowered = head[:512].lower().lstrip()
    if any(lowered.startswith(marker) for marker in _HTML_MARKERS):
        return (
            "this is a web page saved with a spreadsheet's name - several "
            "Indian banks' 'Excel' download does exactly this. Open it in "
            "Excel or your browser, then Save As CSV, and it will read."
        )

    if suffix in WORKBOOK_SUFFIXES:
        return (
            "the name says .xlsx but the contents are not a workbook. If it "
            "is really a CSV, rename it to .csv and it will read."
        )
    return ""


# ------------------------------------------------------------------ values


def render(value: Any) -> str:
    """One spreadsheet cell as the string a CSV would have held.

    The point of this function is that everything after it is the code that
    already works. The reader, the profiler, the format-settling and the
    money parser were all written against a merchant's CSV, and none of them
    should learn what a serial date is. So a workbook becomes strings here and
    is a CSV from this line on.

    Three conversions are load-bearing:

    **Dates.** Excel stores a date as a number of days since 1900 and a
    display format, and openpyxl has already turned that back into a
    `datetime`. Written out as ISO, which the format-settling step then
    recognises without a question - a workbook, unusually, is the one source
    that cannot be ambiguous about day and month.

    **Whole floats.** Every number in a spreadsheet is a float, including
    order ids and reference numbers. Writing `12345.0` where the file shows
    `12345` would turn every identifier into one that matches nothing.

    **Booleans before integers.** `bool` is a subclass of `int` in Python, so
    the obvious order silently renders `TRUE` as `1`.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, dt.datetime):
        if (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0):
            return value.date().isoformat()
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value or value in (float("inf"), float("-inf")):  # NaN / infinity
            return ""
        if value.is_integer():
            return str(int(value))
        # `repr` gives the shortest string that round-trips, which for a
        # number somebody typed is exactly what they typed. Only when that
        # comes back long - the signature of Excel having done arithmetic - is
        # it worth cutting, and then to a fixed width rather than a guess.
        exact = repr(value)
        fraction = exact.partition(".")[2]
        if len(fraction) <= MAX_FRACTION_DIGITS and "e" not in exact:
            return exact
        return f"{value:.{MAX_FRACTION_DIGITS}f}".rstrip("0").rstrip(".")
    return str(value).strip()


# ------------------------------------------------------------------ sheets


def sheets(path: Path) -> tuple[tuple[str, list[list[str]]], ...]:
    """Every sheet in a workbook, as a title and a grid of strings.

    All of them, not the first. A gateway export is routinely one workbook
    with the settlement rows on one sheet and the payments on another, and a
    reader that took `wb.active` would import half a month and report no
    error - the single worst failure this package can have, because the
    arithmetic downstream still balances, just over less money than the
    merchant actually took.

    Sheets Excel has hidden are read too. A hidden sheet is usually a working
    column or a pivot cache and will fail to place, which costs nothing; a
    hidden sheet that holds the data is rare but real, and skipping it would
    lose it silently.
    """
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        # `data_only` reads the value Excel last cached for a formula rather
        # than the formula text. A workbook written by a program and never
        # opened in Excel has no cached values, so a computed column arrives
        # empty - which the profiler reports as a column that fits nothing,
        # and is far better than importing the literal string `=D2*0.02`.
        book = load_workbook(path, read_only=True, data_only=True)
    except InvalidFileException as failure:
        raise FormatError(
            f"{path.name} is not a workbook openpyxl can read: {failure}"
        ) from failure
    except (OSError, ValueError, KeyError, zipfile.BadZipFile) as failure:
        raise FormatError(f"{path.name} could not be opened as a workbook: {failure}") from failure

    found: list[tuple[str, list[list[str]]]] = []
    try:
        for sheet in book.worksheets:
            grid = [[render(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
            # Trailing empty rows and columns are what a spreadsheet's used
            # range looks like after somebody has deleted a few rows, and they
            # arrive as hundreds of blank lines that the header finder would
            # have to wade through.
            while grid and not any(cell.strip() for cell in grid[-1]):
                grid.pop()
            if grid:
                found.append((str(sheet.title), _trim_columns(grid)))
    finally:
        book.close()

    if not found:
        raise FormatError(f"{path.name} has no sheet with anything in it")
    return tuple(found)


def _trim_columns(grid: list[list[str]]) -> list[list[str]]:
    """Drop trailing columns that are empty in every row, and square the grid.

    A sheet's used range is generous on the right in the same way it is at the
    bottom. Left-hand and interior blank columns are kept: a blank column
    between two groups of data is a real thing in a real export, and removing
    it would shift every column heading one place left of its values.
    """
    width = max((len(row) for row in grid), default=0)
    while width > 0 and all(width > len(row) or not row[width - 1].strip() for row in grid):
        width -= 1
    return [[*row[:width], *([""] * (width - len(row)))] for row in grid]
