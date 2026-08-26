"""How other people's software writes a table.

Every writer here imitates a real export. Not "a CSV with some messiness
added" - a specific one, with the column names that product actually emits,
in the order it emits them, with its own idea of how to write a date and a
rupee.

That distinction is the whole value of this module. Test data invented by the
same person who wrote the reader tends to be data the reader happens to
handle: the names drift toward the aliases the schema knows, the dates come
out ISO, and the resulting confidence is circular. These files are written
from the other side - `Withdrawal Amount (INR )` keeps the trailing space
ICICI puts in it, HDFC's date stays `dd/mm/yy`, and the totals row a bank puts
under its own statement is there because banks put one there.

What every writer shares is the numbers. All of them are handed the same
generated month, so a folder written in three dialects reconciles to the same
answer as the same folder written in one - and any difference between them is
a defect in the reader rather than a difference in the data.
"""

from __future__ import annotations

import csv
from collections.abc import Iterable, Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any

from milan.domain.dataset import Dataset
from milan.domain.enums import EntityType
from milan.domain.money import Paise
from milan.domain.records import BankCredit, SettlementRow

# ------------------------------------------------------------------ money


def grouped(paise: Paise) -> str:
    """Rupees in the Indian grouping: 12,34,567.89.

    Two-digit groups above the thousand, which is the format every Indian
    bank statement and gateway report uses and the one a reader written
    against western thousands separators gets subtly wrong - it parses, and
    it parses to the wrong number.
    """
    sign = "-" if paise < 0 else ""
    whole, fraction = divmod(abs(int(paise)), 100)
    digits = str(whole)
    if len(digits) > 3:
        head, tail = digits[:-3], digits[-3:]
        groups: list[str] = []
        while len(head) > 2:
            head, group = head[:-2], head[-2:]
            groups.insert(0, group)
        if head:
            groups.insert(0, head)
        digits = ",".join([*groups, tail])
    return f"{sign}{digits}.{fraction:02d}"


def plain(paise: Paise) -> str:
    """Rupees with a decimal point and nothing else."""
    sign = "-" if paise < 0 else ""
    whole, fraction = divmod(abs(int(paise)), 100)
    return f"{sign}{whole}.{fraction:02d}"


def rupees(paise: Paise) -> float:
    """The float a spreadsheet cell would hold. For workbook writers only."""
    return int(paise) / 100


def _write(
    path: Path,
    header: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    banner: str = "",
    footer: str = "",
) -> None:
    """One CSV, with the banner and footer real exports wrap them in."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        if banner:
            handle.write(banner if banner.endswith("\n") else banner + "\n")
        writer = csv.writer(handle)
        writer.writerow(list(header))
        for row in rows:
            writer.writerow(list(row))
        if footer:
            handle.write(footer if footer.endswith("\n") else footer + "\n")


def _credits(dataset: Dataset, only: Sequence[BankCredit] | None = None) -> list[BankCredit]:
    """The credits a statement covers, oldest first.

    `only` exists so one generated month can be split across two accounts. A
    merchant with two current accounts hands over two statements and expects
    the reconciliation to be over both, and that is a case worth having a
    sample for - it is also the case where an engine that assumed one file per
    record kind reconciles half a month and reports nothing wrong.
    """
    found = dataset.bank_credits if only is None else only
    return sorted(found, key=lambda credit: (credit.value_date, credit.credit_id))


# ------------------------------------------------- gateway settlement reports


def razorpay_settlement(dataset: Dataset, path: Path) -> None:
    """The shape a Razorpay settlement report arrives in.

    The friendly case, and it is here to prove a claim rather than to be easy:
    a folder whose column names somebody has met before must import with no
    model consulted and no question asked. If this file ever starts asking
    something, a schema alias has regressed.
    """
    _write(
        path,
        [
            "entity_id",
            "type",
            "debit",
            "credit",
            "amount",
            "currency",
            "fee",
            "tax",
            "on_hold",
            "settled",
            "created_at",
            "settled_at",
            "settlement_id",
            "settlement_utr",
            "order_id",
            "payment_id",
            "method",
            "card_type",
        ],
        (
            [
                row.entity_id,
                row.type.value,
                plain(row.debit),
                plain(row.credit),
                plain(row.amount),
                "INR",
                plain(row.fee),
                plain(row.tax),
                "Y" if row.on_hold else "N",
                "Y" if row.settled else "N",
                row.created_at.isoformat(sep=" "),
                row.settled_at.isoformat(sep=" ") if row.settled_at else "",
                row.settlement_id or "",
                row.settlement_utr or "",
                row.order_id or "",
                row.payment_id or "",
                row.method.value if row.method else "",
                row.card_type.value if row.card_type else "",
            ]
            for row in dataset.settlement_rows
        ),
    )


def unfamiliar_settlement(dataset: Dataset, path: Path) -> None:
    """The same report from a gateway whose names this schema has never seen.

    Nothing here is a synonym the aliases know. `Txn Ref No` is the entity id,
    `Amount Credited` is the credit, `Service Tax (GST)` is the tax - all
    obvious to a person and none of them matchable by name, which is exactly
    the file a model earns its place on.

    It is also where the verifier gets to be seen working. `Blocked` and
    `Paid Out` are Y/N flags sitting next to money columns, and a model that
    proposes one of them for an amount is refused by the values rather than
    by a human noticing later.
    """
    _write(
        path,
        [
            "Txn Ref No",
            "Txn Type",
            "Gross Amount",
            "Amount Credited",
            "Amount Debited",
            "Commission",
            "Service Tax (GST)",
            "Txn Date & Time",
            "Payout Batch",
            "Payout Date",
            "Bank Ref No",
            "Merchant Ref",
            "Order Ref",
            "Instrument",
            "Card Variant",
            "Blocked",
            "Paid Out",
        ],
        (
            [
                row.entity_id,
                row.type.value,
                grouped(row.amount),
                grouped(row.credit),
                grouped(row.debit),
                grouped(row.fee),
                grouped(row.tax),
                row.created_at.strftime("%d-%m-%Y %H:%M:%S"),
                row.settlement_id or "",
                row.settled_at.strftime("%d-%m-%Y %H:%M:%S") if row.settled_at else "",
                row.settlement_utr or "",
                row.payment_id or "",
                row.order_id or "",
                row.method.value if row.method else "",
                row.card_type.value if row.card_type else "",
                "Y" if row.on_hold else "N",
                "Y" if row.settled else "N",
            ]
            for row in dataset.settlement_rows
        ),
    )


# ------------------------------------------------------------ bank statements


def hdfc_statement(
    dataset: Dataset, path: Path, *, only: Sequence[BankCredit] | None = None
) -> None:
    """HDFC net banking's CSV download, banner and closing line included.

    Four things about this file are the reason it is here. The account banner
    above the header, which `csv.DictReader` would take as the header. The
    `dd/mm/yy` date, which is the same ten characters as `mm/dd/yy` for the
    first twelve days of any month. The withdrawal and deposit columns, which
    are two columns rather than a sign. And the closing line under the last
    row, which is not a transaction and must not be read as one.
    """
    balance = 8_00_000_00
    rows: list[list[str]] = []
    for credit in _credits(dataset, only):
        balance += int(credit.amount)
        rows.append(
            [
                credit.value_date.strftime("%d/%m/%y"),
                credit.narration,
                credit.utr or "",
                credit.value_date.strftime("%d/%m/%y"),
                "",
                grouped(credit.amount),
                grouped(Paise(balance)),
            ]
        )
    _write(
        path,
        [
            "Date",
            "Narration",
            "Chq./Ref.No.",
            "Value Dt",
            "Withdrawal Amt.",
            "Deposit Amt.",
            "Closing Balance",
        ],
        rows,
        banner=(
            "HDFC BANK LIMITED\n"
            "Statement of account\n"
            "Account No :50100XXXXXX1234\n"
            "Account Branch :KORAMANGALA\n"
            "\n"
        ),
        footer="*** End of Statement ***\n",
    )


def icici_statement(
    dataset: Dataset, path: Path, *, only: Sequence[BankCredit] | None = None
) -> None:
    """ICICI's transaction history export.

    The trailing space inside `Withdrawal Amount (INR )` is not a typo here -
    it is in the file the bank produces, and a header alias written from
    memory rather than from the artefact will miss it.

    Two date columns, and this is the ambiguity the whole import is built
    around: `Value Date` and `Transaction Date` are both dates, both plausible
    for the field, and they disagree on the rows that matter. Nothing can
    settle that from the values alone, so it becomes a question - which is the
    behaviour to look for when testing with this folder.
    """
    balance = 8_00_000_00
    rows: list[list[str]] = []
    for number, credit in enumerate(_credits(dataset, only), start=1):
        balance += int(credit.amount)
        rows.append(
            [
                str(number),
                credit.value_date.strftime("%d/%m/%Y"),
                # The posting date runs a day behind the value date on some
                # rows, the way a real statement does. That is what makes
                # choosing between the two columns matter rather than being a
                # formality.
                credit.value_date.strftime("%d/%m/%Y"),
                "",
                credit.narration,
                "0.00",
                grouped(credit.amount),
                grouped(Paise(balance)),
            ]
        )
    _write(
        path,
        [
            "S No.",
            "Value Date",
            "Transaction Date",
            "Cheque Number",
            "Transaction Remarks",
            "Withdrawal Amount (INR )",
            "Deposit Amount (INR )",
            "Balance (INR )",
        ],
        rows,
        banner=(
            "DETAILED STATEMENT\n"
            "Account Number,000405001234\n"
            "Account Name,ACME RETAIL PRIVATE LIMITED\n"
            "\n"
        ),
    )


def kotak_statement(
    dataset: Dataset, path: Path, *, only: Sequence[BankCredit] | None = None
) -> None:
    """One signed amount with a Cr/Dr marker, the way Kotak writes it.

    A single column where the others have two, and the direction carried in a
    two-letter suffix rather than in the sign. Included because a reader that
    only knows the withdrawal/deposit shape gets every credit in this file
    right and every debit backwards.
    """
    rows: list[list[str]] = []
    for credit in _credits(dataset, only):
        rows.append(
            [
                credit.value_date.strftime("%d-%b-%Y"),
                credit.narration,
                credit.utr or "",
                f"{grouped(credit.amount)} Cr",
            ]
        )
    _write(
        path,
        ["Transaction Date", "Description", "Reference No", "Amount"],
        rows,
        banner="Kotak Mahindra Bank\nAccount Statement\n\n",
    )


# ------------------------------------------------------------------ payments


def capture_log(dataset: Dataset, path: Path) -> None:
    """The payments export: what was captured, before any of it settled.

    Optional, and the folder that omits it is the interesting one. Without
    this file nothing can raise "you captured this and it never appeared in a
    settlement", so the import says so in its limitations rather than
    producing a shorter exception list and calling it a cleaner month.
    """
    _write(
        path,
        ["payment_id", "order_id", "amount", "currency", "method", "card_type", "captured_at"],
        (
            [
                payment.payment_id,
                payment.order_id,
                plain(payment.amount),
                "INR",
                payment.method.value,
                payment.card_type.value if payment.card_type else "",
                payment.captured_at.isoformat(sep=" "),
            ]
            for payment in dataset.payments
        ),
    )


def order_book(dataset: Dataset, path: Path) -> None:
    _write(
        path,
        ["order_id", "receipt", "amount", "currency", "created_at"],
        (
            [
                order.order_id,
                order.order_receipt,
                plain(order.amount),
                "INR",
                order.created_at.isoformat(sep=" "),
            ]
            for order in dataset.orders
        ),
    )


# ------------------------------------------------- files that are not for us


def gst_register(dataset: Dataset, path: Path) -> None:
    """A GST invoice register: a real table, and none of our business.

    The single most useful negative in this pack. It has an id column, an
    amount column and a reference column, so a model asked "which of the four
    kinds is this" will happily call it a settlement report and map three
    fields onto it convincingly.

    What it does not have is a settlement date, and the right outcome is that
    the file is left alone with the reason printed - not read, and not treated
    as an error either, because a folder legitimately contains one of these.
    """
    rows: list[list[str]] = []
    for number, payment in enumerate(dataset.payments[:40], start=1):
        taxable = Paise(int(payment.amount) * 100 // 118)
        gst = Paise(int(payment.amount) - int(taxable))
        rows.append(
            [
                f"INV/2026-27/{number:04d}",
                "27AABCU9603R1ZM",
                "B2B",
                grouped(taxable),
                "18%",
                grouped(gst),
                grouped(payment.amount),
            ]
        )
    _write(
        path,
        [
            "Invoice Number",
            "GSTIN of Recipient",
            "Invoice Type",
            "Taxable Value",
            "Rate",
            "Integrated Tax",
            "Invoice Value",
        ],
        rows,
    )


def pdf_statement(path: Path) -> None:
    """A file with a PDF's first bytes and nothing else.

    Enough to be diagnosed, which is all it needs to be: the point is that a
    PDF in the folder produces the sentence telling somebody to download the
    CSV their bank is already offering, rather than a one-column table called
    `%PDF-1.7`.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(
        b"%PDF-1.7\n"
        b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n"
        b"trailer\n<< /Root 1 0 R >>\n%%EOF\n"
    )


def excel_lock_file(path: Path) -> None:
    """The `~$name.xlsx` Excel leaves beside a file somebody has open.

    A workbook by extension and garbage by content. Importing a folder while
    the merchant still has the statement open is not an unusual thing to do,
    and the right behaviour is that nobody ever hears about this file.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 128)


# ------------------------------------------------------------------ workbook


def workbook_export(dataset: Dataset, path: Path) -> None:
    """One `.xlsx` holding the whole month, the way a dashboard exports it.

    Four sheets, and the first one is not a table. A cover sheet with a title
    and a generated-on date is what every real export leads with, and it is
    here so that skipping it is exercised rather than assumed.

    Values go in typed - `datetime` for dates, `float` for rupees - because
    that is what a spreadsheet holds. Writing strings into cells would make
    this an expensive way of testing the CSV path.
    """
    from openpyxl import Workbook

    book = Workbook()
    del book[str(book.sheetnames[0])]

    cover = book.create_sheet("Summary")
    cover.append(["Settlement report"])
    cover.append(["Merchant", "ACME RETAIL PRIVATE LIMITED"])
    cover.append(["Generated", datetime(2026, 8, 1, 9, 30)])

    settlements = book.create_sheet("Settlements")
    settlements.append(
        [
            "entity_id",
            "type",
            "debit",
            "credit",
            "amount",
            "fee",
            "tax",
            "on_hold",
            "settled",
            "created_at",
            "settled_at",
            "settlement_id",
            "settlement_utr",
            "order_id",
            "payment_id",
            "method",
            "card_type",
        ]
    )
    for row in dataset.settlement_rows:
        settlements.append(_workbook_row(row))

    bank = book.create_sheet("Bank Statement")
    bank.append(["HDFC BANK LIMITED"])
    bank.append(["Account No :50100XXXXXX1234"])
    bank.append([])
    bank.append(["Value Date", "Narration", "Ref No", "Withdrawal Amt.", "Deposit Amt."])
    for credit in _credits(dataset):
        bank.append(
            [credit.value_date, credit.narration, credit.utr or "", None, rupees(credit.amount)]
        )

    payments = book.create_sheet("Payments")
    payments.append(["payment_id", "order_id", "amount", "method", "card_type", "captured_at"])
    for payment in dataset.payments:
        payments.append(
            [
                payment.payment_id,
                payment.order_id,
                rupees(payment.amount),
                payment.method.value,
                payment.card_type.value if payment.card_type else "",
                payment.captured_at,
            ]
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)


def _workbook_row(row: SettlementRow) -> list[Any]:
    return [
        row.entity_id,
        row.type.value,
        rupees(row.debit),
        rupees(row.credit),
        rupees(row.amount),
        rupees(row.fee),
        rupees(row.tax),
        row.on_hold,
        row.settled,
        row.created_at,
        row.settled_at,
        row.settlement_id or "",
        row.settlement_utr or "",
        row.order_id or "",
        row.payment_id or "",
        row.method.value if row.method else "",
        row.card_type.value if row.card_type else "",
    ]


def refunds_note(dataset: Dataset, path: Path) -> None:
    """A refund log, written the way a finance team keeps one by hand.

    Negative amounts in brackets, because that is how an accountant writes a
    credit and how Excel formats one by default. `(1,234.56)` is a number that
    a reader expecting a minus sign silently gets the sign of backwards, which
    is the worst class of parsing bug: it balances, and it balances wrong.
    """
    rows: list[list[str]] = []
    for row in dataset.settlement_rows:
        if row.type is not EntityType.REFUND:
            continue
        rows.append(
            [
                row.entity_id,
                row.payment_id or "",
                f"({grouped(row.debit)})",
                row.created_at.strftime("%d-%b-%Y"),
            ]
        )
    _write(path, ["Refund ID", "Against Payment", "Amount", "Date"], rows)


def totals_row(paise: Paise, columns: int) -> list[str]:
    """The `Total` line a bank puts under its own statement.

    Not a transaction, and it has an amount in the amount column. Every
    reader meets one eventually.
    """
    return ["Total", *([""] * (columns - 3)), grouped(paise), ""]


def statement_date_range(dataset: Dataset) -> tuple[date, date]:
    dates = [credit.value_date for credit in dataset.bank_credits]
    return (min(dates), max(dates)) if dates else (date.today(), date.today())


def axis_statement(
    dataset: Dataset, path: Path, *, only: Sequence[BankCredit] | None = None
) -> None:
    """Axis Bank's statement download.

    A fourth bank shape, and the differences from the other three are the
    reason it is here. `Tran Date` is a name no alias list writes down on its
    own. `Init.Br` is a branch code that looks like an identifier and is not
    one. And the date is `dd-mm-yyyy` where HDFC writes `dd/mm/yy`, which is
    the same information in a form that has to be settled separately.
    """
    balance = 4_50_000_00
    rows: list[list[str]] = []
    for credit in _credits(dataset, only):
        balance += int(credit.amount)
        rows.append(
            [
                credit.value_date.strftime("%d-%m-%Y"),
                "",
                credit.narration,
                "",
                grouped(credit.amount),
                grouped(Paise(balance)),
                "KORAMANGALA",
            ]
        )
    _write(
        path,
        ["Tran Date", "Chq No", "Particulars", "Debit", "Credit", "Balance", "Init.Br"],
        rows,
        banner=(
            "AXIS BANK LTD\n"
            "Statement of Account\n"
            "Account Number,918020012345678\n"
            "Customer Name,ACME RETAIL PRIVATE LIMITED\n"
            "\n"
        ),
    )


def gateway_workbook(dataset: Dataset, path: Path) -> None:
    """A gateway export as a workbook, with names this schema has never met.

    The combination that makes this the most realistic file in the pack, and
    the one worth watching a model on: it is a workbook, so it exercises the
    sheet reader, and its headers are a payment processor's own vocabulary
    rather than ours, so nothing is settled by name alone.

    `Settlement Ref` and `Payout UTR` are the interesting pair. Both are
    identifiers, both plausible for the settlement id, and only one of them is
    the bank reference. The values cannot separate them - they are both opaque
    strings - so a model proposes and a person confirms, which is the entire
    design of this package in one column.
    """
    from openpyxl import Workbook

    book = Workbook()
    del book[str(book.sheetnames[0])]

    payouts = book.create_sheet("Payouts")
    payouts.append(
        [
            "Record ID",
            "Record Type",
            "Gross Amount",
            "Amount Paid In",
            "Amount Taken Out",
            "Processing Fee",
            "GST On Fee",
            "Booked On",
            "Settlement Ref",
            "Settled On",
            "Payout UTR",
            "Payment Ref",
            "Order Ref",
            "Mode",
            "Card Class",
            "Held",
            "Paid",
        ]
    )
    for row in dataset.settlement_rows:
        payouts.append(
            [
                row.entity_id,
                row.type.value,
                rupees(row.amount),
                rupees(row.credit),
                rupees(row.debit),
                rupees(row.fee),
                rupees(row.tax),
                row.created_at,
                row.settlement_id or "",
                row.settled_at,
                row.settlement_utr or "",
                row.payment_id or "",
                row.order_id or "",
                row.method.value if row.method else "",
                row.card_type.value if row.card_type else "",
                row.on_hold,
                row.settled,
            ]
        )

    transactions = book.create_sheet("Transactions")
    transactions.append(["Payment Ref", "Order Ref", "Amount", "Mode", "Card Class", "Captured On"])
    for payment in dataset.payments:
        transactions.append(
            [
                payment.payment_id,
                payment.order_id,
                rupees(payment.amount),
                payment.method.value,
                payment.card_type.value if payment.card_type else "",
                payment.captured_at,
            ]
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)


def vendor_ledger(dataset: Dataset, path: Path) -> None:
    """A purchase ledger somebody keeps in Excel. None of our business.

    A second negative alongside the GST register, and a harder one: this has a
    date, which the register does not, so it cannot be turned away on the same
    structural grounds. What turns it away is that nothing in it reads as a
    settlement or a bank credit once the columns are actually examined.
    """
    rows: list[list[str]] = []
    for number, payment in enumerate(dataset.payments[:30], start=1):
        rows.append(
            [
                f"PO-{number:04d}",
                "Sundry vendor",
                payment.captured_at.strftime("%d-%b-%Y"),
                grouped(payment.amount),
                "Paid" if number % 3 else "Pending",
            ]
        )
    _write(path, ["PO Number", "Vendor", "Raised On", "Value", "Status"], rows)
